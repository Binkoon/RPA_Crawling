# Developer : 디지털전략팀 / 강현빈 사원
# Date : 2025/06/26
# 선사링크 : https://ss.shipmentlink.com/tvs2/jsp/TVS2_VesselSchedule.jsp
# 선박 리스트 : ["EVER LUCID","EVER ELITE","EVER LASTING","EVER VIM"]
# 추가 정보 : 하나의 tr에  ARR , DEP이 같이 있음. 따라서 엑셀 전처리 작업이 추가로 필요함.

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options

from .base import ParentsClass
import pandas as pd
import time, os
import logging
import traceback

import openpyxl

class EVERGREEN_Crawling(ParentsClass):
    def __init__(self):
        super().__init__()
        self.carrier_name = "EMC"
        
        # 로깅 설정
        self.setup_logging()
        
        # 선박 리스트
        self.vessel_name_list = ["EVER LUCID","EVER ELITE","EVER LASTING","EVER VIM"]
        
        # 크롤링 결과 추적
        self.success_count = 0
        self.fail_count = 0
        self.failed_vessels = []

    def setup_logging(self):
        """로깅 설정"""
        # 초기에는 에러가 없으므로 파일 로그 생성하지 않음
        self.logger = super().setup_logging(self.carrier_name, has_error=False)
        
    def setup_logging_with_error(self):
        """에러 발생 시 로깅 설정"""
        # 에러가 발생했으므로 파일 로그 생성
        self.logger = super().setup_logging(self.carrier_name, has_error=True)

    def step1_visit_website_and_handle_cookie(self):
        """1단계: 선사 홈페이지 접속하는데 쿠키허용 팝업이 뜨면 처리해주고, 없으면 다음 스텝으로 이동"""
        try:
            self.logger.info("=== 1단계: 선사 홈페이지 접속 및 쿠키 처리 시작 ===")
            
            # 0. 선사 홈페이지 접속
            self.Visit_Link("https://ss.shipmentlink.com/tvs2/jsp/TVS2_VesselSchedule.jsp")
            driver = self.driver
            wait = self.wait

            # 쿠키 뜨는 경우 //*[@id="btn_cookie_accept_all"]
            try:
                cookie_btn = wait.until(EC.element_to_be_clickable((
                    By.XPATH , '//*[@id="btn_cookie_accept_all"]'
                )))
                cookie_btn.click()
                self.logger.info("쿠키 허용 팝업 처리 완료")
            except:
                self.logger.info("쿠키 팝업이 없거나 이미 처리됨")
            
            self.logger.info("=== 1단계: 선사 홈페이지 접속 및 쿠키 처리 완료 ===")
            return True
            
        except Exception as e:
            # 에러 발생 시 로깅 설정 변경
            self.setup_logging_with_error()
            self.logger.error(f"=== 1단계: 선사 홈페이지 접속 및 쿠키 처리 실패 ===")
            self.logger.error(f"에러 메시지: {str(e)}")
            self.logger.error(f"상세 에러: {traceback.format_exc()}")
            return False

    def step2_crawl_vessel_data(self):
        """2단계: 지정된 선박별로 루핑 시작"""
        try:
            self.logger.info("=== 2단계: 선박별 데이터 크롤링 시작 ===")
            
            driver = self.driver
            wait = self.wait

            for vessel_name in self.vessel_name_list:
                try:
                    self.logger.info(f"선박 {vessel_name} 크롤링 시작")
                    
                    vessel_select_elem = wait.until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="vslCode"]'))
                    )
                    vessel_select = Select(vessel_select_elem)

                    found = False
                    for option in vessel_select.options:
                        if vessel_name in option.text:
                            vessel_select.select_by_visible_text(option.text)
                            self.logger.info(f"선박명 '{vessel_name}' 선택 완료")
                            found = True
                            break

                    if not found:
                        self.logger.warning(f"선박 {vessel_name}을 찾을 수 없습니다.")
                        self.fail_count += 1
                        self.failed_vessels.append(vessel_name)
                        continue

                    time.sleep(1)

                    submit_btn = wait.until(
                        EC.element_to_be_clickable((By.XPATH, '//*[@id="submitButton"]'))
                    )
                    submit_btn.click()
                    self.logger.info(f"Submit 버튼 클릭 완료")

                    time.sleep(2)  # 데이터 로딩 대기

                    # 선박조회 후 쿠키 팝업이 나타날 수 있으므로 처리
                    try:
                        cookie_btn = wait.until(EC.element_to_be_clickable((
                            By.XPATH, '//*[@id="btn_cookie_accept_all"]'
                        )))
                        cookie_btn.click()
                        self.logger.info("선박조회 후 쿠키 허용 팝업 처리 완료")
                    except:
                        self.logger.info("선박조회 후 쿠키 팝업이 없거나 이미 처리됨")  

                    # ====== 항차번호(span[text]) 추출 ======
                    # span[1], span[3], span[5], ... 순회
                    table_titles = []
                    span_idx = 1
                    while True:
                        try:
                            span_xpath = f'//*[@id="schedule"]/table/tbody/tr[2]/td/table/tbody/tr[2]/td/table/tbody/tr[2]/td[2]/span[{span_idx}]'
                            span_elem = driver.find_element(By.XPATH, span_xpath)
                            table_titles.append(span_elem.text.strip())
                            span_idx += 2
                        except:
                            break  # 더 이상 span이 없으면 종료

                    # ====== 테이블 데이터 수집 ======
                    all_tables = []
                    table_idx = 1
                    while True:
                        try:
                            # 각 테이블의 1, 2번째 row만 추출
                            table_data = []
                            for row_idx in [1, 2]:
                                row_xpath = f'//*[@id="schedule"]/table/tbody/tr[2]/td/table/tbody/tr[2]/td/table/tbody/tr[2]/td[2]/table[{table_idx}]/tbody/tr/td/table/tbody/tr[{row_idx}]'
                                row_elem = wait.until(
                                    EC.presence_of_element_located((By.XPATH, row_xpath))
                                )
                                tds = row_elem.find_elements(By.TAG_NAME, "td")
                                row_data = [td.text.strip() for td in tds]
                                table_data.append(row_data)

                            # 테이블 데이터 + 항차번호 함께 저장
                            voyage = table_titles[table_idx - 1] if table_idx - 1 < len(table_titles) else ""
                            all_tables.append({"voyage": voyage, "table": table_data})
                            table_idx += 1
                        except Exception:
                            # 더 이상 table이 없으면 break
                            break

                    # ====== 데이터 저장 ======
                    # 테이블별로 DataFrame 변환 후 concat
                    if all_tables:
                        df_list = []
                        for item in all_tables:
                            table = item["table"]
                            voyage = item["voyage"]
                            if len(table) == 2:
                                df = pd.DataFrame([table[1]], columns=table[0])
                                df["항차번호"] = voyage
                                df_list.append(df)
                        if df_list:
                            result_df = pd.concat(df_list, ignore_index=True)
                            save_path = self.get_save_path(self.carrier_name, vessel_name)
                            result_df.to_excel(save_path, index=False)
                            self.logger.info(f"엑셀 저장 완료: {save_path}")

                            # openpyxl로 wrapText 활성화 (헤더는 제외)
                            wb = openpyxl.load_workbook(save_path)
                            ws = wb.active
                            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                                for cell in row:
                                    cell.alignment = openpyxl.styles.Alignment(wrap_text=True)
                            wb.save(save_path)
                            self.logger.info("텍스트 줄바꿈(wrapText) 활성화 완료")
                            
                            self.success_count += 1
                        else:
                            self.logger.warning(f"선박 {vessel_name}에 대한 데이터가 없습니다.")
                            self.fail_count += 1
                            self.failed_vessels.append(vessel_name)
                    else:
                        self.logger.warning(f"선박 {vessel_name}에 대한 테이블 데이터가 없습니다.")
                        self.fail_count += 1
                        self.failed_vessels.append(vessel_name)
                    
                    self.logger.info(f"선박 {vessel_name} 크롤링 완료")
                    
                except Exception as e:
                    self.logger.error(f"선박 {vessel_name} 크롤링 실패: {str(e)}")
                    self.fail_count += 1
                    self.failed_vessels.append(vessel_name)
                    continue
            
            self.logger.info("=== 2단계: 선박별 데이터 크롤링 완료 ===")
            self.logger.info(f"성공: {self.success_count}개, 실패: {self.fail_count}개")
            return True
            
        except Exception as e:
            # 에러 발생 시 로깅 설정 변경
            self.setup_logging_with_error()
            self.logger.error(f"=== 2단계: 선박별 데이터 크롤링 실패 ===")
            self.logger.error(f"에러 메시지: {str(e)}")
            self.logger.error(f"상세 에러: {traceback.format_exc()}")
            return False

    def step3_save_with_naming_rules(self):
        """3단계: 파일명 규칙 및 저장경로 규칙 적용"""
        try:
            self.logger.info("=== 3단계: 파일명 규칙 및 저장경로 규칙 적용 시작 ===")
            
            # TODO: 파일명 규칙과 저장 경로 규칙 적용 로직 구현 필요
            # 현재는 기존 change_filename 로직을 활용하되, 향후 개선 예정
            for vessel_name in self.vessel_name_list:
                try:
                    self.change_filename(vessel_name)
                    self.logger.info(f"선박 {vessel_name} 파일명 변경 완료")
                except Exception as e:
                    self.logger.error(f"선박 {vessel_name} 파일명 변경 실패: {str(e)}")
                    continue
            
            self.logger.info("=== 3단계: 파일명 규칙 및 저장경로 규칙 적용 완료 ===")
            return True
            
        except Exception as e:
            # 에러 발생 시 로깅 설정 변경
            self.setup_logging_with_error()
            self.logger.error(f"=== 3단계: 파일명 규칙 및 저장경로 규칙 적용 실패 ===")
            self.logger.error(f"에러 메시지: {str(e)}")
            self.logger.error(f"상세 에러: {traceback.format_exc()}")
            return False

    def step4_preprocessing_logic(self):
        """4단계: 전처리 로직 수행"""
        try:
            self.logger.info("=== 4단계: 전처리 로직 수행 시작 ===")
            
            # ARR/DEP 데이터 분리 로직
            for vessel_name in self.vessel_name_list:
                try:
                    file_path = self.get_save_path(self.carrier_name, vessel_name)
                    if os.path.exists(file_path):
                        # 엑셀 파일 읽기
                        df = pd.read_excel(file_path)
                        
                        # ARR/DEP 컬럼이 있는지 확인
                        if 'ARR/DEP' in df.columns:
                            # ARR/DEP 컬럼을 ARR과 DEP로 분리
                            df[['ARR', 'DEP']] = df['ARR/DEP'].str.split('/', expand=True)
                            df = df.drop('ARR/DEP', axis=1)
                            
                            # 변경된 데이터를 다시 저장
                            df.to_excel(file_path, index=False)
                            self.logger.info(f"선박 {vessel_name} ARR/DEP 분리 완료")
                        else:
                            self.logger.info(f"선박 {vessel_name} ARR/DEP 컬럼 없음")
                    else:
                        self.logger.warning(f"선박 {vessel_name} 파일이 존재하지 않음")
                        
                except Exception as e:
                    self.logger.error(f"선박 {vessel_name} 전처리 실패: {str(e)}")
                    continue
            
            self.logger.info("=== 4단계: 전처리 로직 수행 완료 ===")
            return True
            
        except Exception as e:
            # 에러 발생 시 로깅 설정 변경
            self.setup_logging_with_error()
            self.logger.error(f"=== 4단계: 전처리 로직 수행 실패 ===")
            self.logger.error(f"에러 메시지: {str(e)}")
            self.logger.error(f"상세 에러: {traceback.format_exc()}")
            return False

    def run(self):
        """메인 실행 함수"""
        try:
            self.logger.info("=== EVERGREEN 크롤링 시작 ===")
            
            # 1단계: 선사 홈페이지 접속 및 쿠키 처리
            if not self.step1_visit_website_and_handle_cookie():
                return False
            
            # 2단계: 지정된 선박별로 루핑 작업 시작
            if not self.step2_crawl_vessel_data():
                return False
            
            # 3단계: 파일명 규칙 및 저장경로 규칙 적용
            if not self.step3_save_with_naming_rules():
                return False
            
            # 4단계: 전처리 로직 수행
            if not self.step4_preprocessing_logic():
                return False
            
            # 최종 결과 로깅
            self.logger.info("=== EVERGREEN 크롤링 완료 ===")
            self.logger.info(f"총 {len(self.vessel_name_list)}개 선박 중")
            self.logger.info(f"성공: {self.success_count}개")
            self.logger.info(f"실패: {self.fail_count}개")
            if self.failed_vessels:
                self.logger.info(f"실패한 선박: {', '.join(self.failed_vessels)}")
            
            self.Close()
            return True
            
        except Exception as e:
            # 에러 발생 시 로깅 설정 변경
            self.setup_logging_with_error()
            self.logger.error(f"=== EVERGREEN 크롤링 전체 실패 ===")
            self.logger.error(f"에러 메시지: {str(e)}")
            self.logger.error(f"상세 에러: {traceback.format_exc()}")
            return False
